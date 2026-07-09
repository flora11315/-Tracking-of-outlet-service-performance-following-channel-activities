from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import json

import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
INPUT_XLSX = BASE / "上海片区渠道活动后网点服务效果追踪Demo模拟数据.xlsx"
OUTPUT_DIR = BASE / "skill_outputs"
CHART_DIR = OUTPUT_DIR / "weekly_charts"
REPORT_XLSX = OUTPUT_DIR / "channel_activity_service_report.xlsx"


@dataclass
class NodeAlert:
    alert_id: str
    stage: str
    priority: str
    score: float
    cadence: str
    wecom_group: str
    event_id: str
    event_date: str
    channel_name: str
    channel_level: str
    bank_type: str
    event_size: str
    theme: str
    participants: int
    holding_signal: str
    signal: str
    material_need: str
    material_source: str
    suggested_action: str
    due_days: int
    owner_action_required: str


def read_sheet(sheet_name: str, required: bool = True) -> pd.DataFrame:
    try:
        df = pd.read_excel(INPUT_XLSX, sheet_name=sheet_name, header=3)
    except ValueError:
        if required:
            raise
        return pd.DataFrame()
    df = df.dropna(how="all")
    if "活动ID" in df.columns:
        df = df[df["活动ID"].notna()].copy()
    return df


def to_number(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    for col in cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def infer_material_from_theme(theme: str) -> str:
    if "指数" in theme:
        return "指数投资讲稿"
    if "养老" in theme:
        return "养老投教折页"
    if "波动" in theme:
        return "波动市场客户问答"
    if "权益" in theme or "风险收益" in theme or "固收" in theme:
        return "产品风险收益解释页"
    if "市场展望" in theme or "资产配置" in theme:
        return "月度市场观点材料"
    return "产品风险收益解释页"


def infer_material(row: pd.Series) -> tuple[str, str]:
    survey_need = row.get("后续材料需求")
    if pd.notna(survey_need) and str(survey_need).strip():
        return str(survey_need).strip(), "问卷反馈"

    t30 = float(row.get("T+30保有量变化率_%", 0) or 0)
    t90 = float(row.get("T+90保有量变化率_%", 0) or 0)
    profit = float(row.get("T+90盈利占比_%", 100) or 100)
    redemption = float(row.get("赎回压力_%", 0) or 0)
    if redemption >= 85 or t30 < 0:
        return "波动市场客户问答", "经营信号推断"
    if profit < 50 or t90 < 0:
        return "产品风险收益解释页", "经营信号推断"
    return infer_material_from_theme(str(row.get("活动主题", ""))), "活动主题映射"


def derive_service_label(row: pd.Series) -> str:
    t90 = float(row.get("T+90保有量变化率_%", 0) or 0)
    profit = float(row.get("T+90盈利占比_%", 50) or 50)
    redemption = float(row.get("赎回压力_%", 0) or 0)
    activity_score = row.get("活动质量评分")
    view_score = row.get("投教观点质量评分")

    if t90 >= 1.5 and profit >= 50 and redemption < 80:
        return "稳定提升"
    if t90 < 0 or profit < 48 or redemption >= 95:
        return "需加强陪伴"
    if pd.notna(activity_score) and pd.notna(view_score) and (activity_score < 4.5 or view_score < 4.5):
        return "内容/陪伴优化"
    return "观察跟踪"


def build_analysis_table() -> pd.DataFrame:
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Input workbook not found: {INPUT_XLSX}")

    activities = read_sheet("02_活动表")
    tracking = read_sheet("03_经营追踪")
    survey = read_sheet("05_问卷反馈", required=False)

    activity_cols = [
        "活动ID",
        "活动日期",
        "活动层级",
        "渠道ID",
        "渠道名称",
        "银行类型",
        "活动规模类型",
        "活动主题",
        "实际参与人数",
        "渠道销售/服务团队",
    ]
    tracking_cols = [
        "活动ID",
        "活动前保有量_万元",
        "T+30保有量_万元",
        "T+90保有量_万元",
        "T+30申购_万元",
        "T+30赎回_万元",
        "T+90累计申购_万元",
        "T+90累计赎回_万元",
        "T+30保有量变化率_%",
        "T+90保有量变化率_%",
        "T+90盈利占比_%",
        "赎回压力_%",
    ]
    df = activities[activity_cols].merge(tracking[tracking_cols], on="活动ID", how="left")

    if not survey.empty:
        survey_cols = [
            "活动ID",
            "活动质量评分",
            "投教观点质量评分",
            "后续材料需求",
            "后续服务建议",
            "渠道反馈摘要",
        ]
        df = df.merge(survey[survey_cols], on="活动ID", how="left")
    else:
        df["活动质量评分"] = pd.NA
        df["投教观点质量评分"] = pd.NA
        df["后续材料需求"] = pd.NA
        df["后续服务建议"] = pd.NA
        df["渠道反馈摘要"] = pd.NA

    df = to_number(
        df,
        [
            "实际参与人数",
            "活动前保有量_万元",
            "T+30保有量_万元",
            "T+90保有量_万元",
            "T+30申购_万元",
            "T+30赎回_万元",
            "T+90累计申购_万元",
            "T+90累计赎回_万元",
            "T+30保有量变化率_%",
            "T+90保有量变化率_%",
            "T+90盈利占比_%",
            "赎回压力_%",
            "活动质量评分",
            "投教观点质量评分",
        ],
    )

    material_results = df.apply(infer_material, axis=1)
    df["材料需求"] = [result[0] for result in material_results]
    df["材料需求来源"] = [result[1] for result in material_results]
    df["服务效果标签"] = df.apply(derive_service_label, axis=1)
    df["建议输出"] = df.apply(base_recommendation, axis=1)
    return df


def base_recommendation(row: pd.Series) -> str:
    material = row.get("材料需求", "产品风险收益解释页")
    source = row.get("材料需求来源", "经营信号推断")
    label = derive_service_label(row)
    if label == "稳定提升":
        return f"沉淀为优秀案例，后续延续《{material}》和同类活动模板。"
    if label == "需加强陪伴":
        return f"补充《{material}》，安排短会或二次陪伴；材料需求来源：{source}。"
    if label == "内容/陪伴优化":
        return f"复盘活动主题和材料表达，补充《{material}》；材料需求来源：{source}。"
    return f"常规跟踪，补充《{material}》；材料需求来源：{source}。"


def infer_wecom_group(row: pd.Series) -> str:
    bank_type = str(row["银行类型"])
    channel_level = str(row["活动层级"])
    if channel_level == "分行":
        return f"上海片区-{bank_type}-分行服务群"
    if channel_level == "支行":
        return f"上海片区-{bank_type}-支行服务群"
    return f"上海片区-{bank_type}-网点陪伴群"


def priority_for_stage(stage: str, row: pd.Series) -> tuple[str, int, str]:
    activity_score = row.get("活动质量评分")
    view_score = row.get("投教观点质量评分")
    t30 = float(row.get("T+30保有量变化率_%", 0) or 0)
    t90 = float(row.get("T+90保有量变化率_%", 0) or 0)
    profit = float(row.get("T+90盈利占比_%", 50) or 50)
    redemption = float(row.get("赎回压力_%", 0) or 0)
    label = str(row.get("服务效果标签", "观察跟踪"))

    if stage == "T+7":
        if pd.notna(activity_score) and pd.notna(view_score) and (activity_score < 4.3 or view_score < 4.3):
            return "高", 3, "需负责人确认"
        if pd.notna(activity_score) and pd.notna(view_score) and (activity_score < 4.8 or view_score < 4.8):
            return "中", 5, "建议跟进"
        return "低", 10, "常规记录"
    if stage == "T+30":
        if t30 < 0 or redemption >= 90:
            return "高", 3, "需负责人确认"
        if redemption >= 75 or t30 < 0.5:
            return "中", 5, "建议跟进"
        return "低", 10, "常规记录"
    if "需加强陪伴" in label or t90 < 0 or profit < 48:
        return "高", 3, "需负责人确认"
    if "内容/陪伴优化" in label or profit < 52:
        return "中", 5, "建议跟进"
    return "低", 10, "常规记录"


def holding_signal(row: pd.Series) -> str:
    return (
        f"T+30 {row['T+30保有量变化率_%']:.2f}%，"
        f"T+90 {row['T+90保有量变化率_%']:.2f}%，"
        f"赎回压力 {row['赎回压力_%']:.1f}%"
    )


def short_signal(stage: str, row: pd.Series) -> str:
    if stage == "T+7":
        if pd.notna(row.get("投教观点质量评分")) and row["投教观点质量评分"] < 4.5:
            return "投教反馈需优化"
        if pd.notna(row.get("活动质量评分")) and row["活动质量评分"] < 4.5:
            return "活动体验需优化"
        return f"材料需求：{row['材料需求']}"
    if stage == "T+30":
        if float(row["赎回压力_%"]) >= 75:
            return f"赎回压力偏高，{holding_signal(row)}"
        if float(row["T+30保有量变化率_%"]) < 0:
            return f"保有量短期波动，{holding_signal(row)}"
        return f"短期经营信号平稳，{holding_signal(row)}"
    if row["服务效果标签"] == "稳定提升":
        return f"优秀案例，{holding_signal(row)}"
    if row["服务效果标签"] == "需加强陪伴":
        return f"需持续陪伴，{holding_signal(row)}"
    return f"{row['服务效果标签']}，{holding_signal(row)}"


def short_action(stage: str, row: pd.Series, priority: str) -> str:
    material = row["材料需求"]
    if stage == "T+7":
        return f"补充《{material}》" + (" + 复盘材料" if priority == "高" else "")
    if stage == "T+30":
        return f"安排15分钟短会 + 补充《{material}》" if priority == "高" else f"补充《{material}》并继续观察"
    if row["服务效果标签"] == "稳定提升":
        return "沉淀为优秀案例"
    if priority == "高":
        return f"纳入持续服务名单 + 补充《{material}》"
    return "进入阶段复盘"


def attention_score(stage: str, row: pd.Series, priority: str) -> float:
    score = {"高": 60, "中": 35, "低": 10}[priority]
    score += {"T+7": 4, "T+30": 8, "T+90": 6}[stage]
    score += {"分行": 8, "支行": 5, "网点": 2}.get(str(row["活动层级"]), 0)
    if float(row["T+30保有量变化率_%"]) < 0:
        score += 10
    if float(row["T+90保有量变化率_%"]) < 0:
        score += 10
    if float(row["赎回压力_%"]) >= 90:
        score += 12
    elif float(row["赎回压力_%"]) >= 75:
        score += 7
    if float(row["T+90盈利占比_%"]) < 48:
        score += 10
    elif float(row["T+90盈利占比_%"]) < 52:
        score += 5
    activity_score = row.get("活动质量评分")
    view_score = row.get("投教观点质量评分")
    if pd.notna(activity_score) and activity_score < 4.5:
        score += 5
    if pd.notna(view_score) and view_score < 4.5:
        score += 5
    return score


def excellence_score(row: pd.Series) -> float:
    score = 0.0
    score += max(float(row["T+90保有量变化率_%"]), 0) * 8
    score += float(row["T+90盈利占比_%"]) / 5
    score += max(100 - float(row["赎回压力_%"]), 0) / 10
    if pd.notna(row.get("活动质量评分")):
        score += float(row["活动质量评分"]) * 2
    if pd.notna(row.get("投教观点质量评分")):
        score += float(row["投教观点质量评分"]) * 2
    if row["活动规模类型"] == "小场":
        score += 8
    if row["服务效果标签"] == "稳定提升":
        score += 12
    return score


def make_alerts(df: pd.DataFrame) -> list[NodeAlert]:
    alerts: list[NodeAlert] = []
    for _, row in df.iterrows():
        for stage in ["T+7", "T+30", "T+90"]:
            priority, due_days, owner_action = priority_for_stage(stage, row)
            alerts.append(
                NodeAlert(
                    alert_id=f"ALT-{len(alerts) + 1:03d}",
                    stage=stage,
                    priority=priority,
                    score=attention_score(stage, row, priority),
                    cadence=f"节点触发：活动后第 {stage[2:]} 天",
                    wecom_group=infer_wecom_group(row),
                    event_id=str(row["活动ID"]),
                    event_date=str(row["活动日期"]),
                    channel_name=str(row["渠道名称"]),
                    channel_level=str(row["活动层级"]),
                    bank_type=str(row["银行类型"]),
                    event_size=str(row["活动规模类型"]),
                    theme=str(row["活动主题"]),
                    participants=int(row["实际参与人数"]),
                    holding_signal=holding_signal(row),
                    signal=short_signal(stage, row),
                    material_need=str(row["材料需求"]),
                    material_source=str(row["材料需求来源"]),
                    suggested_action=short_action(stage, row, priority),
                    due_days=due_days,
                    owner_action_required=owner_action,
                )
            )
    return alerts


def alert_to_dict(alert: NodeAlert) -> dict:
    return {
        "提醒ID": alert.alert_id,
        "节点": alert.stage,
        "优先级": alert.priority,
        "关注分": round(alert.score, 1),
        "提醒机制": "企业微信机器人节点提醒",
        "提醒频率": alert.cadence,
        "企微目标群": alert.wecom_group,
        "活动ID": alert.event_id,
        "活动日期": alert.event_date,
        "渠道名称": alert.channel_name,
        "活动层级": alert.channel_level,
        "银行类型": alert.bank_type,
        "活动规模类型": alert.event_size,
        "活动主题": alert.theme,
        "实际参与人数": alert.participants,
        "保有量观察": alert.holding_signal,
        "提醒信号": alert.signal,
        "材料需求": alert.material_need,
        "材料需求来源": alert.material_source,
        "建议动作": alert.suggested_action,
        "建议完成时限_工作日": alert.due_days,
        "负责人动作": alert.owner_action_required,
        "是否进入邮件附件": "是",
        "处理状态": "未处理",
    }


def priority_sort_value(priority: str) -> int:
    return {"高": 0, "中": 1, "低": 2}.get(priority, 3)


def stage_sort_value(stage: str) -> int:
    return {"T+7": 0, "T+30": 1, "T+90": 2}.get(stage, 3)


def wecom_markdown(alert: NodeAlert) -> str:
    color = {"高": "warning", "中": "comment", "低": "info"}.get(alert.priority, "comment")
    return f"""### 【{alert.stage}｜<font color=\"{color}\">{alert.priority}</font>】{alert.channel_name}
> 活动：{alert.theme}｜{alert.event_size}｜{alert.participants}人
> 信号：{alert.signal}
> 保有量：{alert.holding_signal}
> 动作：{alert.suggested_action}
> 来源：{alert.material_source}｜时限：{alert.due_days}个工作日
"""


def write_wecom_messages(alerts: list[NodeAlert]) -> None:
    ordered = sorted(alerts, key=lambda a: (stage_sort_value(a.stage), priority_sort_value(a.priority), -a.score))
    lines = [
        "# 企业微信机器人节点提醒文案",
        "",
        "说明：短版企业微信提醒。材料需求若非问卷得出，会在“来源”中标注为经营信号推断或活动主题映射。",
        "",
    ]
    payloads = []
    for alert in ordered:
        content = wecom_markdown(alert)
        lines.append(f"## {alert.alert_id}｜{alert.stage}｜{alert.priority}优先级｜{alert.channel_name}")
        lines.append("")
        lines.append(content)
        lines.append("")
        payloads.append(
            {
                "alert_id": alert.alert_id,
                "stage": alert.stage,
                "priority": alert.priority,
                "target_group": alert.wecom_group,
                "payload": {"msgtype": "markdown", "markdown": {"content": content}},
            }
        )
    (OUTPUT_DIR / "wecom_node_messages.md").write_text("\n".join(lines), encoding="utf-8")
    (OUTPUT_DIR / "wecom_node_payloads.json").write_text(json.dumps(payloads, ensure_ascii=False, indent=2), encoding="utf-8")


def chart_color(label: str) -> str:
    return {
        "T+7": "#4C78A8",
        "T+30": "#F58518",
        "T+90": "#54A24B",
        "高": "#D95550",
        "中": "#F2B447",
        "低": "#5DA5DA",
    }.get(label, "#4C78A8")


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for candidate in ["/System/Library/Fonts/PingFang.ttc", "/System/Library/Fonts/STHeiti Light.ttc"]:
        try:
            return ImageFont.truetype(candidate, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    value = hex_color.lstrip("#")
    return tuple(int(value[i : i + 2], 16) for i in (0, 2, 4))


def write_bar_chart(title: str, data: dict[str, int], path: Path, color_by_label: bool = False) -> None:
    scale = 2
    width = 760
    row_h = 38
    top = 64
    left = 185
    right = 36
    bottom = 28
    height = top + bottom + row_h * max(len(data), 1)
    max_value = max(data.values()) if data else 1
    bar_max = width - left - right - 90
    img = Image.new("RGB", (width * scale, height * scale), "white")
    draw = ImageDraw.Draw(img)
    title_font = load_font(20 * scale)
    body_font = load_font(14 * scale)
    draw.text((24 * scale, 17 * scale), title, font=title_font, fill=hex_to_rgb("#0B2545"))
    for idx, (label, value) in enumerate(data.items()):
        y = top + idx * row_h
        bar_w = int(bar_max * value / max_value) if max_value else 0
        color = chart_color(label) if color_by_label else "#4C78A8"
        draw.text((24 * scale, (y + 8) * scale), str(label), font=body_font, fill=hex_to_rgb("#1F2937"))
        draw.rounded_rectangle(
            [left * scale, (y + 6) * scale, (left + bar_w) * scale, (y + 26) * scale],
            radius=4 * scale,
            fill=hex_to_rgb(color),
        )
        draw.text(((left + bar_w + 10) * scale, (y + 8) * scale), str(value), font=body_font, fill=hex_to_rgb("#374151"))
    img.save(path)


def write_charts(df: pd.DataFrame, alerts: list[NodeAlert]) -> dict[str, Path]:
    CHART_DIR.mkdir(parents=True, exist_ok=True)
    alert_df = pd.DataFrame([alert_to_dict(a) for a in alerts])
    stage_counts = alert_df["节点"].value_counts().reindex(["T+7", "T+30", "T+90"]).fillna(0).astype(int).to_dict()
    priority_counts = alert_df["优先级"].value_counts().reindex(["高", "中", "低"]).fillna(0).astype(int).to_dict()
    material_counts = df["材料需求"].value_counts().head(5).astype(int).to_dict()
    paths = {
        "stage": CHART_DIR / "stage_distribution.png",
        "priority": CHART_DIR / "priority_distribution.png",
        "material": CHART_DIR / "material_needs_top5.png",
    }
    write_bar_chart("节点提醒分布", stage_counts, paths["stage"], color_by_label=True)
    write_bar_chart("优先级分布", priority_counts, paths["priority"], color_by_label=True)
    write_bar_chart("高频材料需求 Top 5", material_counts, paths["material"])
    return paths


def top_focus(alerts: list[NodeAlert], limit: int = 5) -> list[NodeAlert]:
    selected: dict[str, NodeAlert] = {}
    for alert in sorted(alerts, key=lambda a: (-a.score, stage_sort_value(a.stage), a.event_id)):
        if alert.event_id not in selected:
            selected[alert.event_id] = alert
        if len(selected) >= limit:
            break
    return list(selected.values())


def top_excellent_cases(df: pd.DataFrame, limit: int = 3) -> pd.DataFrame:
    cases = df.copy()
    cases["优秀案例分"] = cases.apply(excellence_score, axis=1)
    cases = cases[(cases["服务效果标签"] == "稳定提升") & (cases["活动规模类型"] == "小场")]
    return cases.sort_values("优秀案例分", ascending=False).head(limit)


def focus_line(alert: NodeAlert) -> str:
    return f"{alert.channel_name}：{alert.stage} {alert.signal}，建议{alert.suggested_action}。"


def case_line(row: pd.Series) -> str:
    return (
        f"{row['渠道名称']}{row['活动规模类型']}：{row['活动主题']}，"
        f"T+90 保有量 {row['T+90保有量变化率_%']:.2f}%，盈利占比 {row['T+90盈利占比_%']:.1f}%，"
        "适合沉淀为优秀案例。"
    )


def next_week_actions(alerts: list[NodeAlert]) -> list[str]:
    high = [a for a in alerts if a.priority == "高"]
    top_material = pd.Series([a.material_need for a in high]).value_counts().index[0] if high else "产品风险收益解释页"
    return [
        f"向高优先级渠道补充《{top_material}》。",
        "对 T+30 赎回压力较高的渠道安排 15 分钟线上短会。",
        "沉淀 T+90 稳定提升的小场活动标准模板。",
    ]


def write_email_digest(df: pd.DataFrame, alerts: list[NodeAlert], focus: list[NodeAlert], excellent: pd.DataFrame) -> None:
    alert_df = pd.DataFrame([alert_to_dict(a) for a in alerts])
    material_top = "、".join(df["材料需求"].value_counts().head(3).index.tolist())
    focus_count = len({a.event_id for a in alerts if a.priority == "高"})
    reusable_count = len(excellent)
    lines = [
        "主题：上海片区渠道活动后服务效果周报",
        "",
        "各位同事好，",
        "",
        "## 一句话结论",
        f"本期 {len(df)} 场活动均已完成 T+7/T+30/T+90 节点追踪，其中 {focus_count} 个渠道建议重点陪伴，{reusable_count} 个小场案例可沉淀复用。",
        "",
        "## 本期概览",
        f"- 覆盖活动：{len(df)} 场",
        f"- 节点提醒：{len(alerts)} 条",
        f"- 高优先级：{int((alert_df['优先级'] == '高').sum())} 条",
        f"- 高频材料需求：{material_top}",
        "",
        "## 重点关注",
    ]
    for idx, alert in enumerate(focus, start=1):
        lines.append(f"{idx}. {focus_line(alert)}")
    lines.extend(["", "## 优秀案例"])
    for idx, (_, row) in enumerate(excellent.iterrows(), start=1):
        lines.append(f"{idx}. {case_line(row)}")
    lines.extend(["", "## 下周建议动作"])
    for action in next_week_actions(alerts):
        lines.append(f"- {action}")
    lines.extend(["", "附件：channel_activity_service_report.xlsx（含节点提醒明细、分析宽表、数据概览和图表数据）"])
    lines.extend(["", "以上为系统基于渠道聚合数据生成的服务提醒，请结合实际渠道沟通情况判断后续动作。"])
    (OUTPUT_DIR / "scheduled_email_digest.md").write_text("\n".join(lines), encoding="utf-8")


def write_report_workbook(df: pd.DataFrame, alerts: list[NodeAlert], focus: list[NodeAlert], excellent: pd.DataFrame) -> None:
    alert_df = pd.DataFrame([alert_to_dict(a) for a in alerts])
    material_stats = df["材料需求"].value_counts().reset_index()
    material_stats.columns = ["材料需求", "活动数"]
    material_source_stats = df["材料需求来源"].value_counts().reset_index()
    material_source_stats.columns = ["材料需求来源", "活动数"]
    stage_stats = alert_df["节点"].value_counts().reindex(["T+7", "T+30", "T+90"]).fillna(0).reset_index()
    stage_stats.columns = ["节点", "提醒数"]
    priority_stats = alert_df["优先级"].value_counts().reindex(["高", "中", "低"]).fillna(0).reset_index()
    priority_stats.columns = ["优先级", "提醒数"]
    size_label = pd.crosstab(df["活动规模类型"], df["服务效果标签"]).reset_index()

    summary = pd.DataFrame(
        [
            ["覆盖活动", len(df)],
            ["节点提醒", len(alerts)],
            ["高优先级提醒", int((alert_df["优先级"] == "高").sum())],
            ["材料需求Top1", material_stats.iloc[0, 0] if not material_stats.empty else ""],
            ["说明", "06_分析宽表由活动表、经营追踪表和可选问卷反馈自动生成，不作为原始输入。"],
        ],
        columns=["项目", "内容"],
    )
    focus_df = pd.DataFrame([alert_to_dict(a) for a in focus])
    excellent_df = excellent.copy()

    with pd.ExcelWriter(REPORT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="00_邮件摘要", index=False)
        alert_df.to_excel(writer, sheet_name="01_节点提醒明细", index=False)
        pd.concat(
            {
                "节点分布": stage_stats,
                "优先级分布": priority_stats,
                "活动规模与标签": size_label,
            },
            names=["分析模块", "行号"],
        ).reset_index(level=0).to_excel(writer, sheet_name="02_活动汇总分析", index=False)
        focus_df.to_excel(writer, sheet_name="03_重点关注Top5", index=False)
        excellent_df.to_excel(writer, sheet_name="04_优秀案例Top3", index=False)
        material_stats.to_excel(writer, sheet_name="05_材料需求统计", index=False)
        start_row = 0
        for title, table in [
            ("节点提醒分布", stage_stats),
            ("优先级分布", priority_stats),
            ("材料需求Top5", material_stats.head(5)),
            ("材料来源分布", material_source_stats),
        ]:
            pd.DataFrame([[title]], columns=["图表数据"]).to_excel(writer, sheet_name="06_图表数据", index=False, startrow=start_row)
            table.to_excel(writer, sheet_name="06_图表数据", index=False, startrow=start_row + 2)
            start_row += len(table) + 5
        df.to_excel(writer, sheet_name="07_分析宽表", index=False)

    format_report_workbook()


def format_report_workbook() -> None:
    wb = load_workbook(REPORT_XLSX)
    header_fill = PatternFill("solid", fgColor="E8EEF5")
    title_fill = PatternFill("solid", fgColor="0B2545")
    border = Border(bottom=Side(style="thin", color="D9E2EC"))
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="0B2545")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 32)

    chart_ws = wb["06_图表数据"]
    # Add simple native charts in the data overview sheet.
    chart1 = BarChart()
    chart1.title = "节点提醒分布"
    data = Reference(chart_ws, min_col=2, min_row=3, max_row=5)
    cats = Reference(chart_ws, min_col=1, min_row=4, max_row=5)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart_ws.add_chart(chart1, "D3")

    chart2 = PieChart()
    chart2.title = "优先级分布"
    data2 = Reference(chart_ws, min_col=2, min_row=10, max_row=12)
    cats2 = Reference(chart_ws, min_col=1, min_row=10, max_row=12)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart_ws.add_chart(chart2, "D18")
    wb.save(REPORT_XLSX)


def clean_old_outputs() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for path in OUTPUT_DIR.iterdir():
        if path.is_file():
            path.unlink()
        elif path.is_dir() and path.name == "weekly_charts":
            for child in path.iterdir():
                if child.is_file():
                    child.unlink()


def main() -> None:
    clean_old_outputs()
    df = build_analysis_table()
    alerts = make_alerts(df)
    focus = top_focus(alerts, 5)
    excellent = top_excellent_cases(df, 3)
    alert_df = pd.DataFrame([alert_to_dict(a) for a in alerts])
    alert_df.to_csv(OUTPUT_DIR / "service_node_alerts.csv", index=False, encoding="utf-8-sig")
    df.to_csv(OUTPUT_DIR / "generated_analysis_table.csv", index=False, encoding="utf-8-sig")
    write_wecom_messages(alerts)
    write_charts(df, alerts)
    write_email_digest(df, alerts, focus, excellent)
    write_report_workbook(df, alerts, focus, excellent)

    summary = {
        "input_workbook": str(INPUT_XLSX),
        "input_sheets": ["02_活动表", "03_经营追踪", "05_问卷反馈(可选)"],
        "events_read": int(len(df)),
        "node_alerts_generated": int(len(alerts)),
        "stage_counts": alert_df["节点"].value_counts().to_dict(),
        "priority_counts": alert_df["优先级"].value_counts().to_dict(),
        "material_source_counts": df["材料需求来源"].value_counts().to_dict(),
        "outputs": [
            "scheduled_email_digest.md",
            "channel_activity_service_report.xlsx",
            "service_node_alerts.csv",
            "generated_analysis_table.csv",
            "wecom_node_messages.md",
            "wecom_node_payloads.json",
        ],
    }
    (OUTPUT_DIR / "run_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
